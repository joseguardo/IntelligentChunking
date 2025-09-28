# The goal of this modules is to create a pipeline that: 
# 1. Takes an excel template file as input and sends it to the OpenAI API one sheet at a time.
# 2. Divides it into the different sheets that it contains. 
# 3. Processes each sheet in order to find all of the fields of the template that need to be completed in the sheet. 
# 4. For each field it will generate an entry in a BaseModel where it will discribe the name of the field (which must be the same as in the excel tempplate), brief description, and the type of data that the template requires for the field. 
# 5. Finally it will return a list of BaseModels, one for each sheet in the excel template.


# GOAL: Automatically transform excel templates into structured data models that allow autonomized search and retrieval of information from filled templates.

import os
import re
import tempfile
import asyncio
import json
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional
from textwrap import dedent
from pprint import pprint
from mistralai import Mistral


from pydantic import BaseModel, Field
from openpyxl import load_workbook, Workbook
from dotenv import load_dotenv
from openai import OpenAI  # pip install openai
from llama_cloud_services import LlamaParse  # pip install llama-cloud-services (per docs)

# Module for querying our generated vector store. 
from vectorStore import WeaviateAsyncManager

# ---------- Your Pydantic models ----------
class ExcelField(BaseModel):
    name: Optional[str] = Field(None, description="The name of the block of fields as it appears in the excel template.")
    fields: Optional[List[str]] = Field(None, description="Fields that are part of the block. ")
    prompt: Optional[str] = Field(None, description="A prompt that can be used to extract the fields value from the contract and that contains context from the sheet and other fields present.")

class ListExcelFields(BaseModel):
    excelfields: List[ExcelField] = Field(..., description="A list of fields extracted from the excel template.")

# ---------- Helpers ----------
SAFE_SHEET_CHARS = r'[\\/*?:\[\]]'
def _sanitize_sheet_name(name: str, maxlen: int = 31) -> str:
    n = re.sub(SAFE_SHEET_CHARS, "_", (name or "").strip())
    return n[:maxlen] or "Sheet"

def _write_single_sheet_tempfile(src_ws) -> str:
    """
    Create a new xlsx on disk containing only the provided worksheet's VALUES (no styles).
    Returns path to the temp file.
    """
    tmp = tempfile.NamedTemporaryFile(prefix="sheet_", suffix=".xlsx", delete=False)
    tmp_path = tmp.name
    tmp.close()

    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = _sanitize_sheet_name(src_ws.title)

    for row in src_ws.iter_rows(values_only=True):
        ws_out.append(row)

    # (optional) column widths for readability (best-effort)
    try:
        for dim, col_dim in src_ws.column_dimensions.items():
            if col_dim.width:
                ws_out.column_dimensions[dim].width = col_dim.width
    except Exception:
        pass

    wb_out.save(tmp_path)
    return tmp_path

# ---------- Main parser ----------
class ExcelParser:
    @staticmethod
    def split_listexcelfields(parsed: 'ListExcelFields') -> list:
        """
        Given a ListExcelFields object, return a list of ListExcelFields,
        each containing a single ExcelField from the original.
        """
        return [ListExcelFields(excelfields=[field]) for field in (parsed.excelfields or [])]
    
    def save_results_to_json(self, results: Dict[str, ListExcelFields], output_path: str = None) -> str:
        """
        Save the parsed results to a JSON file for easy reuse.
        
        Args:
            results: The parsed results from parse_sheets()
            output_path: Optional custom path. If not provided, uses excel filename + '_parsed.json'
            
        Returns:
            The path to the saved JSON file
        """
        if output_path is None:
            excel_name = Path(self.excel_file_path).stem
            output_path = f"{excel_name}_parsed.json"
        
        # Convert Pydantic models to dict for JSON serialization
        json_data = {}
        for sheet_name, list_excel_fields in results.items():
            json_data[sheet_name] = list_excel_fields.model_dump()
        
        # Add metadata
        json_data["_metadata"] = {
            "excel_file": self.excel_file_path,
            "parsed_at": datetime.now().isoformat(),
            "total_sheets": len(results),
            "parser_config": {
                "openai_model": self.openai_model,
                "max_concurrency": self.max_concurrency
            }
        }
        
        # Save to file
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(json_data, f, indent=2, ensure_ascii=False)
        
        print(f"✅ Saved parsed results to: {output_path}")
        return output_path
    
    @staticmethod
    def load_results_from_json(json_path: str) -> Dict[str, ListExcelFields]:
        """
        Load previously saved parsing results from JSON.
        
        Args:
            json_path: Path to the saved JSON file
            
        Returns:
            Dictionary mapping sheet names to ListExcelFields objects
        """
        with open(json_path, 'r', encoding='utf-8') as f:
            json_data = json.load(f)
        
        # Remove metadata
        metadata = json_data.pop('_metadata', {})
        print(f"📁 Loading results from: {json_path}")
        if metadata:
            print(f"   Original file: {metadata.get('excel_file', 'Unknown')}")
            print(f"   Parsed at: {metadata.get('parsed_at', 'Unknown')}")
            print(f"   Total sheets: {metadata.get('total_sheets', 'Unknown')}")
        
        # Convert back to Pydantic models
        results = {}
        for sheet_name, sheet_data in json_data.items():
            results[sheet_name] = ListExcelFields(**sheet_data)
        
        return results
    
    def parse_sheets_with_cache(self, cache_path: str = None, force_reparse: bool = False) -> Dict[str, ListExcelFields]:
        """
        Parse sheets with caching support. Will load from cache if available and not forced to reparse.
        
        Args:
            cache_path: Path to cache file. If None, uses excel filename + '_parsed.json'
            force_reparse: If True, ignores cache and reparses everything
            
        Returns:
            Dictionary mapping sheet names to ListExcelFields objects
        """
        if cache_path is None:
            excel_name = Path(self.excel_file_path).stem
            cache_path = f"{excel_name}_parsed.json"
        
        # Check if cache exists and is not forced to reparse
        if not force_reparse and Path(cache_path).exists():
            try:
                print(f"🔍 Found cached results at: {cache_path}")
                return self.load_results_from_json(cache_path)
            except Exception as e:
                print(f"⚠️  Error loading cache: {e}. Proceeding with fresh parsing...")
        
        # Parse fresh results
        print("🚀 Parsing fresh results...")
        results = self.parse_sheets()
        
        # Handle the tuple return from parse_sheets
        if isinstance(results, tuple):
            results, output_dict = results
        
        # Save to cache
        self.save_results_to_json(results, cache_path)
        
        return results
    def __init__(
        self,
        excel_file_path: str,
        *,
        openai_model: Optional[str] = None,       # model for responses.parse
        include_hidden: bool = False,
        only_sheets: Optional[List[str]] = None,
        exclude_sheets: Optional[List[str]] = None,
        max_concurrency: int = 8,  # CHANGE: used to cap concurrent sheet processing

        # LlamaParse config (matches your snippet; override if desired)
        lp_parse_mode: str = "parse_page_with_agent",
        lp_model: str = "openai-gpt-4o-mini",
        lp_high_res_ocr: bool = True,
        lp_adaptive_long_table: bool = True,
        lp_outlined_table_extraction: bool = True,
        lp_output_tables_as_html: bool = True,
        lp_parse_instructions: str = "Parse only financial fields.",
    ):
        """
        We still return Dict[str, ListExcelFields] to keep your public API intact.
        """
        load_dotenv()

        self.excel_file_path = str(Path(excel_file_path).resolve())
        self.workbook = load_workbook(self.excel_file_path, data_only=True)
        self.max_concurrency = max(1, int(max_concurrency))  # CHANGE: store cap

        # OpenAI client for the final structured extraction
        self.client = OpenAI(api_key=os.getenv("OPENAI_API_KEY"))
        self.openai_model = openai_model or os.getenv("OPENAI_MODEL") or "gpt-4.1"

        # LlamaParse client for robust parsing of each sheet temp file
        self.llama_parser = LlamaParse(
            api_key=os.getenv("LLAMAPARSE_API_KEY"),
            parse_mode=lp_parse_mode,
            model=lp_model,
            high_res_ocr=lp_high_res_ocr,
            adaptive_long_table=lp_adaptive_long_table,
            outlined_table_extraction=lp_outlined_table_extraction,
            output_tables_as_HTML=lp_output_tables_as_html,
            user_prompt=lp_parse_instructions,
        )

        # choose sheets
        all_sheet_names = self.workbook.sheetnames
        if only_sheets:
            target = [s for s in all_sheet_names if s in set(only_sheets)]
        else:
            target = list(all_sheet_names)
        if exclude_sheets:
            target = [s for s in target if s not in set(exclude_sheets)]
        if not include_hidden:
            visible = []
            for s in target:
                ws = self.workbook[s]
                if getattr(ws, "sheet_state", "visible") == "hidden":
                    continue
                visible.append(s)
            target = visible
        self.sheets = target

    # ----------------- Public API (sync) -----------------
    def parse_sheets(self) -> Dict[str, ListExcelFields]:
        """
        Sync wrapper: internally runs async LlamaParse per sheet, then OpenAI parse.
        """
        return asyncio.run(self.parse_sheets_async())

    # ----------------- Async implementation -----------------
    async def parse_sheets_async(self) -> Dict[str, ListExcelFields]:
        results: Dict[str, ListExcelFields] = {}
        sem = asyncio.Semaphore(self.max_concurrency)

        # create tasks
        tasks = [self._process_sheet(sheet_name, sem) for sheet_name in self.sheets]

        # print a header so it's visually clear
        print(f"[{datetime.now():%H:%M:%S}] 🚀 Starting {len(tasks)} sheets with max_concurrency={self.max_concurrency}")
        output_dict = {}
        # show results as they COMPLETE (not in submission order)
        for fut in asyncio.as_completed(tasks):
            sheet_name, parsed = await fut
            results[sheet_name] = parsed

            # live, per-sheet preview
            try:
                line = self._preview_line(sheet_name, parsed, getattr(self, "preview_n", 5))  # CHANGE
                print(line)
            except Exception:
                # don't let logging crash the run
                print(f"[{datetime.now():%H:%M:%S}] ✅ {sheet_name}: {len(parsed.excelfields)} fields")

            try:
                parsed = self.split_listexcelfields(parsed)  # CHANGE: test the new static method
                output_dict[sheet_name] = parsed
            except Exception:
                pass

        print(f"[{datetime.now():%H:%M:%S}] 🏁 All sheets done.")
        return results, output_dict


    # CHANGE: new async per-sheet pipeline (guarded by semaphore)
    async def _process_sheet(self, sheet_name: str, sem: asyncio.Semaphore):
        await sem.acquire()
        tmp_path = None
        try:
            ws = self.workbook[sheet_name]
            tmp_path = _write_single_sheet_tempfile(ws)

            # 1) LlamaParse (async)
            parsed_text = await self._llamaparse_sheet_text(tmp_path)
            print(f"[LlamaParse] Parsed '{sheet_name}' ({len(parsed_text)} chars)")
            print(f"[LlamaParse] Sample:\n{parsed_text}...\n")

            # 1.1) Quick block generation (async)
            parsed_text = await self._llamaparse_parsed_structuring(parsed_text)
            print(f"[LlamaParse] Post-processed '{sheet_name}' ({len(parsed_text)} chars)")
            print(f"[LlamaParse] Sample:\n{parsed_text}...\n")

        
            # 2) OpenAI structuring (sync client -> run in thread)
            result = await self._openai_extract_fields_from_text(sheet_name, parsed_text)

            # Normalize
            result = self.normalize_model_response(result)
            orig_len = len(result.excelfields or [])
            orig_fields = [f.name for f in (result.excelfields or []) if getattr(f, "name", None)]

            # 3) OpenAI validation: iterate through fields and look for extracted fields that could be part of a list such as "Contact 1 Name", "Contact 2 Name", etc.
            result = await self._openai_validate_fields(sheet_name, result)

            # Normalize
            result = self.normalize_model_response(result)
            new_fields = [f.name for f in (result.excelfields or []) if getattr(f, "name", None)]
            new_len = len(result.excelfields or [])
            if new_len != orig_len:
                print(f"[{datetime.now():%H:%M:%S}] 🔄 {sheet_name}: Validated fields, reduced {orig_len} → {new_len}")
                print(f"  - Removed fields: {set(orig_fields) - set(new_fields)}")
                print(f"  - Added fields: {set(new_fields) - set(orig_fields)}")


            return sheet_name, result

        except Exception as e:
            print(f"[ERROR] Sheet '{sheet_name}': {e}")
            return sheet_name, ListExcelFields(excelfields=[])
        finally:
            try:
                if tmp_path and os.path.exists(tmp_path):
                    os.remove(tmp_path)
            except Exception:
                pass
            sem.release()

    # ----------------- Internals -----------------
    async def _llamaparse_sheet_text(self, file_path: str) -> str:
        """
        Use LlamaParse to parse a single-sheet XLSX. Return a single text string.
        We concatenate all returned page texts (if any).
        """
        start = datetime.now()
        print(f"[LlamaParse] Parsing: {Path(file_path).name} at {start:%Y-%m-%d %H:%M:%S}")

        # The docs show:
        # result = await parser.aparse("dcf_template.xlsx")
        # llama_parse_documents = result.get_text_documents(split_by_page=True)
        result = await self.llama_parser.aparse(file_path)
        docs = result.get_text_documents(split_by_page=True)

        # You can choose HTML or text—docs[i].text should contain the parsed content.
        # If you set output_tables_as_HTML=True, text may include HTML tables.
        combined = "\n\n".join([d.text for d in docs if getattr(d, "text", None)])
        if not combined.strip():
            print("[LlamaParse] Warning: Empty parse result; falling back to empty text.")
        return combined
    
    async def _llamaparse_parsed_structuring(self, parsed_text: str) -> str:
        """ 
        Quickly reorganize into semantic blocks using OpenAI responses. 
        """
        prompt = dedent("""
            You are given the content of an EXCEL SHEET TEMPLATE that has been parsed into markdown. 
            Your task is to understand the structure of the template and reorganize it into clear sections or blocks.
            Each section should be clearly labeled with its real header that describes its purpose and is present in the template.
            - Don't add any extra content nor punctuation. 
            - Only organize the existing content into sections.
            - Don't any introductory or closing text.
            - If the template contains tables, lists, treat them as part of the section they belong to.
        """).strip()

        try:
            start = datetime.now()
            print(f"[OpenAI] Structuring parsed text at {start:%H:%M:%S}")

            completion = self.client.responses.create(
                model=self.openai_model,
                input=[{
                    "role": "user",
                    "content": [
                        {"type": "input_text", "text": parsed_text},  # safety trim if huge
                        {"type": "input_text", "text": prompt},
                    ],
                }],
            )
            structured = completion.output_text
            print(f"[LlamaParse] Done structuring with {len(structured)} chars.")
            return structured
        except Exception as e:
            print(f"[LlamaParse ERROR] Structuring: {e}")
            return parsed_text  # fallback to original if error occurs

    # CHANGE: split into sync core (runs OpenAI client) ...
    def _openai_extract_fields_from_text_sync(self, sheet_name: str, parsed_text: str) -> ListExcelFields:
        """
        Call OpenAI Responses API to convert the LlamaParse text into your Pydantic structure. (SYNC)
        """
        prompt = dedent("""
            You are given the content of an EXCEL SHEET TEMPLATE that has been parsed into text/HTML by an OCR/structure parser.
            Your task is to identify any explicit fillable blocks of fields in the template.

            For each block you find, return:
            1) The exact block name as it appears in the template.
            2) A list of fields that are part of the block. 
            3) A professionally engineered prompt that captures the context of the block through the purpose of the sheet and enables the extraction of the block value from a contract through a vector store search.
            This prompt must include output examples, rich context knowledge without being too verbose, and any other information that will help the LLM understand what to extract and how to format it, replicating the style of an agent prompt. 

            Important rules:
            - ONLY include fields explicitly present in the sheet.
            - Don't add any extra punctuation or quotes around field names like "", ":", ";" or ''
            - Understand the template structure and differentiate between labels, headers, and actual fields to fill.
            - Do not hallucinate.
            - Ignore generic headers, boilerplate, or instructional text unless it is itself a field to fill.
            - If the parser emitted tables as HTML, read them as if they were the original grid.
            - Interpret correctly fields that may appear as lists or repeated groups and ONLY include them as one field. 
            - Do not include any case specific or sensitive information that may be included in the template as examples.

            Return strictly as the Pydantic model: ListExcelFields.
        """).strip()

        try:
            start = datetime.now()
            print(f"[OpenAI] Structuring fields for '{sheet_name}' at {start:%H:%M:%S}")

            completion = self.client.responses.parse(
                model=self.openai_model,
                input=[{
                    "role": "user",
                    "content": [
                        {"type": "input_text", "text": parsed_text[:200000]},  # safety trim if huge
                        {"type": "input_text", "text": prompt},
                    ],
                }],
                text_format=ListExcelFields,
            )
            structured = completion.output_parsed
            print(f"[OpenAI] Done '{sheet_name}' with {len(structured.excelfields)} fields.")
            return structured

        except Exception as e:
            print(f"[OpenAI ERROR] '{sheet_name}': {e}")
            return ListExcelFields(excelfields=[])
        


    def _openai_validate_fields_sync(self, sheet_name: str, result: ListExcelFields) -> ListExcelFields:
        """ 
        Call the OpenAI Api to validate the fields extracted from the excel sheet. (SYNC)
        We look for patterns in the field names that suggest they are part of a list (e.g., "Contact 1 Name", "Contact 2 Name").
        If we find such patterns, we consolidate them into a single field. Since the ListExcelFields already is formatted as a list, it would be as if we removed duplicates. 
        """

        prompt = dedent(""" 
        You are an expert data structuring assistant specializing in the normalization of extracted form fields from Excel templates.
        You are given a list of extracted fields, each with a name, description, and data type, as parsed from an Excel sheet.
        Your task is to analyze the list of field names and identify patterns that indicate the fields are part of a repeated group or list.
        For example, fields such as "Contact 1 Name", "Contact 2 Name", "Contact 3 Name" should be recognized as instances of a single logical field ("Contact Name") that occurs multiple times.

        Instructions:

        - Review all field names for numeric or sequential patterns (e.g., "Field 1", "Field 2", "Field 3", etc.).
        - For each group of fields that share a common base name and differ only by a number or similar index, consolidate them into a single field definition.
        - In the output, ensure that each logical field appears only once, with a clear indication in the description that it can occur multiple times if applicable.
        - Data type should still only be one of: 'str', 'int', 'float', 'datetime'.
        - If no patterns are found, return the original list unchanged.
        - Ensure the final output is a valid ListExcelFields Pydantic model.
        - Be careful with fields that may contain same words, but are not part of a list (e.g., "Start Date" and "End Date" are distinct and should not be consolidated).
        - Preserve all unique, non-repeated fields as-is.
        - Return the validated and consolidated list strictly as a ListExcelFields Pydantic model, ensuring the output is clean, deduplicated, and semantically accurate.

        Example: 
            Input: 
            ListExcelFields(excelfields=[
            ExcelField(name="Contact 1 Name", fields=["Name of the first contact"], prompt = "..."),
            ExcelField(name="Contact 2 Name", fields=["Name of the second contact"], prompt = "..."),
            ExcelField(name="Contact 1 Email", fields=["Email of the first contact"], prompt = "..."),
            ExcelField(name="Contact 2 Email", fields=["Email of the second contact"], prompt = "..."),
            ExcelField(name="Company Name", fields=["Name of the company"], prompt = "..."),
        ])
        Output: 
                    ListExcelFields(excelfields=[
                ExcelField(
                    name="Contact Name",
                    fields=["Name of each contact (list, one per contact)"],
                    prompt = "Extract the name of each contact from the contract."
                ),
                ExcelField(
                    name="Contact Email",
                    fields=["Email of each contact (list, one per contact)"],
                    prompt = "Extract the email of each contact from the contract."
                ),
                ExcelField(
                    name="Company Name",
                    fields=["Name of the company"],
                    prompt = "Extract the name of the company from the contract."
                ),
            ])
        Do not hallucinate or invent fields. Only consolidate when a clear pattern is present. If no patterns are found, return the original list unchanged.
                    prompt = "Extract the email of each contact from the contract."
                ),
                ExcelField(
                    name="Company Name",
                    fields=["Name of the company"],
                    prompt = "Extract the name of the company from the contract."
                ),
            ])
        Do not hallucinate or invent fields. Only consolidate when a clear pattern is present. If no patterns are found, return the original list unchanged.
        """).strip()
        try:
            start = datetime.now()
            print(f"[OpenAI] Validating fields for '{sheet_name}' at {start:%H:%M:%S}")

            # Convert the ListExcelFields to a string representation for the prompt
            fields_str = repr(result)

            completion = self.client.responses.parse(
                model=self.openai_model,
                input=[{
                    "role": "user",
                    "content": [
                        {"type": "input_text", "text": fields_str+prompt},
                    ],
                }],
                text_format=ListExcelFields,
            )
            validated = completion.output_parsed
            print(f"[OpenAI] Validation done '{sheet_name}' with {len(validated.excelfields)} fields.")
            return validated

        except Exception as e:
            print(f"[OpenAI ERROR] Validation '{sheet_name}': {e}")
            return result  # return original if validation fails
        


    # CHANGE: ... and async wrapper so we can run it concurrently without blocking the loop
    async def _openai_extract_fields_from_text(self, sheet_name: str, parsed_text: str) -> ListExcelFields:
        return await asyncio.to_thread(self._openai_extract_fields_from_text_sync, sheet_name, parsed_text)
    
    async def _openai_validate_fields(self, sheet_name: str, result: ListExcelFields) -> ListExcelFields:
        return await asyncio.to_thread(self._openai_validate_fields_sync, sheet_name, result)
    
    # CHANGE: helper to build a compact preview line
    def _preview_line(self, sheet_name: str, parsed: ListExcelFields, n: int = 5) -> str:
        names = [f.name for f in (parsed.excelfields or []) if getattr(f, "name", None)]
        head = ", ".join(names) if names else "—"
        #more = f" +{len(names)-n}" if len(names) > n else ""
        return f"[{datetime.now():%H:%M:%S}] ✅ {sheet_name}: {len(parsed.excelfields)} fields  →  {head}"
    
    def normalize_model_response(self, response) -> ListExcelFields:
        if isinstance(response, dict):
            return ListExcelFields(**response)
        elif isinstance(response, ListExcelFields):
            return response
        else:
            return ListExcelFields(excelfields=[])
        
    def convert_excelfield_to_listexcelfields(self, field: ExcelField) -> ListExcelFields:
        return ListExcelFields(excelfields=[field])


def main():
    async def run():
        async with WeaviateAsyncManager() as manager:
            load_dotenv()
            mistral_api_key = os.getenv("MISTRAL_API_KEY")
            model = "mistral-large-latest"

            client = Mistral(api_key=mistral_api_key)
            parser = ExcelParser(
                "Caledonia Facility A.xlsx",
                #only_sheets=[load_workbook("Caledonia Facility A.xlsx").sheetnames[0]],
                openai_model="gpt-5-mini-2025-08-07",   # the structuring step
                lp_model="openai-gpt-4o-mini",          # the LlamaParse internal LLM (per your snippet)
                max_concurrency=4,                      # CHANGE: control parallelism here
            )
            loaded_result = parser.load_results_from_json("Caledonia Facility A_parsed.json")
            result = loaded_result.get("Lenders")
            for field in result.excelfields[0:1]:  # test first field only
                print("===============================================================")
                print(f"- {field.name}: {field.fields} \n  Prompt: {field.prompt}\n")
                query = field.prompt    
                results = await manager.async_query(collection_name="ContractClausesWithTenancy", query=query, limit=3, tenant="tenantA")
                print("--------------")
                for result in results:
                    prompt = query + "\n\n" + result["clause_name"] + "\n\n" + result["clause_text"]
                    print(f"Prompt: {prompt}\n\n")
                    print("\n\n")
                    print("Generating response from Mistral...\n\n")                    
                    chat_response = client.chat.complete(
                        model=model,
                        messages=[
                            {
                                "role": "user",
                                "content": prompt
                            }
                        ]
                    )
                    completion = chat_response.choices[0].message.content
                    print(f"Response number: {completion}\n\n")
                    print("--------------\n\n")

    asyncio.run(run())


if __name__ == "__main__":
    # env:
    #   OPENAI_API_KEY=...
    #   LLAMAPARSE_API_KEY=...
    #   (optional) OPENAI_MODEL=gpt-4.1
    
    # Use caching - will load from cache if available, otherwise parse fresh
    #result = parser.parse_sheets_with_cache()
    # Or force reparse: result = parser.parse_sheets_with_cache(force_reparse=True)
    # Or specify custom cache path: result = parser.parse_sheets_with_cache("custom_cache.json")
    main()