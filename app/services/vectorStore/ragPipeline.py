from pydantic import BaseModel, Field
from typing import Optional, List, Dict, Any, Type
from concurrent.futures import ThreadPoolExecutor
import re
from dotenv import load_dotenv
import os 
from excelParser import ExcelParser
from excelExporter import ExcelExporter, export_extraction_results
from openpyxl import load_workbook, Workbook


import os
import re
from typing import Dict, List, Optional, Type, Any
from concurrent.futures import ThreadPoolExecutor
from dotenv import load_dotenv
from pydantic import BaseModel, Field


class ExcelField(BaseModel):
    name: Optional[str] = Field(None, description="The name of the block of fields as it appears in the excel template.")
    fields: Optional[List[str]] = Field(None, description="Fields that are part of the block.")
    prompt: Optional[str] = Field(None, description="A prompt that can be used to extract the fields value from the contract and that contains context from the sheet and other fields present.")


class ListExcelFields(BaseModel):
    excelfields: List[ExcelField] = Field(..., description="A list of fields extracted from the excel template.")


class DocumentProcessor:
    """Main class that handles document processing with dynamic field extraction."""
    
    def __init__(self, client, llm_model: str = "gpt-5-mini-2025-08-07", vector_store_id: Optional[str] = None):
        """
        Initialize the document processor.
        
        Args:
            client: OpenAI client instance
            llm_model: The LLM model to use for extraction
            vector_store_id: Vector store ID for file search
        """
        load_dotenv()  # Load environment variables from .env file
        self.client = client
        self.llm_model = llm_model
        self.vector_store_id = os.getenv("VECTOR_STORE", vector_store_id)

    def convert_name_to_field(self, name: str) -> str:
        """Convert field name to valid Python identifier by replacing spaces with underscores."""
        return re.sub(r'\s+', '_', name.strip()).lower()

    def create_extracted_field_classes(self, list_excel_fields: ListExcelFields) -> Dict[str, Dict[str, Any]]:
        """
        Create dynamic Pydantic BaseModel classes for each ExcelField.
        Each class contains a field based on the name and individual fields from the fields list.
        
        Args:
            list_excel_fields: ListExcelFields object containing the fields to convert
            
        Returns:
            Dictionary mapping field names to their corresponding BaseModel classes and metadata
        """
        created_classes = {}
        
        for excel_field in list_excel_fields.excelfields:
            if not excel_field.name:
                continue
            
            # Convert name to valid Python identifier for class name
            class_name = self.convert_name_to_field(excel_field.name).title().replace('_', '')
            
            # Initialize class dictionary and annotations
            class_annotations = {}
            class_dict = {
                '__module__': __name__, 
                '__annotations__': class_annotations
            }
            
            # Add field based on the name value
            name_field_key = self.convert_name_to_field(excel_field.name)
            class_annotations[name_field_key] = Optional[str]
            class_dict[name_field_key] = Field(
                None, 
                description=f"Field for {excel_field.name}"
            )
            
            # Add all fields from the fields list
            if excel_field.fields:
                for field_str in excel_field.fields:
                    if field_str:  # Skip empty strings
                        field_key = self.convert_name_to_field(field_str)
                        # Avoid duplicate field names
                        if field_key not in class_annotations:
                            class_annotations[field_key] = Optional[str]
                            class_dict[field_key] = Field(
                                None, 
                                description=f"Field for {field_str}"
                            )
            
            # Create the dynamic BaseModel class
            DynamicClass = type(
                class_name,
                (BaseModel,),
                class_dict
            )
            
            # Store only essential data for pipeline functionality
            created_classes[excel_field.name] = {
                'master_class': DynamicClass,
                'prompt': excel_field.prompt or "",
                'data_type': 'str'
            }
        
        return created_classes

    def get_prompts_from_created_classes(self, created_classes: Dict[str, Dict[str, Any]]) -> Dict[str, str]:
        """
        Extract all prompts from the created classes structure.
        
        Args:
            created_classes: Output from create_extracted_field_classes
            
        Returns:
            Dictionary mapping field names to their prompts
        """
        prompts = {}
        for field_name, class_info in created_classes.items():
            prompt = class_info.get('prompt', '').strip()
            if prompt:
                prompts[field_name] = prompt
        return prompts

    def get_prompt_for_field(self, created_classes: Dict[str, Dict[str, Any]], field_name: str) -> str:
        """
        Get prompt for a specific field from created classes.
        
        Args:
            created_classes: Output from create_extracted_field_classes
            field_name: Name of the field to get prompt for
            
        Returns:
            The prompt string for the specified field
            
        Raises:
            ValueError: If field not found or has no prompt
        """
        if field_name not in created_classes:
            raise ValueError(f"Field '{field_name}' not found in created classes.")
        
        prompt = created_classes[field_name].get('prompt', '').strip()
        if not prompt:
            raise ValueError(f"Field '{field_name}' has no prompt.")
        
        return prompt

    def process_dynamic_basemodels(self, list_excel_fields: ListExcelFields) -> Dict[str, Dict[str, Any]]:
        """
        Process dynamic Pydantic BaseModel classes for each ExcelField.
        Creates dynamic classes and performs parallel extraction.
        
        Args:
            list_excel_fields: ListExcelFields object containing the fields to convert
            
        Returns:
            Dictionary mapping field names to their extraction results
        """
        print(f"Processing {len(list_excel_fields.excelfields)} fields...")
        
        # 1. Create dynamic Pydantic classes for each field
        created_classes = self.create_extracted_field_classes(list_excel_fields)
        print(f"Created {len(created_classes)} dynamic classes")
        
        # 2. Extract prompts for each field
        field_prompts = self.get_prompts_from_created_classes(created_classes)
        print(f"Extracted prompts for {len(field_prompts)} fields")
        
        if not field_prompts:
            print("No fields with prompts found. Nothing to extract.")
            return {}
        
        # 3. Create extraction tasks for each field type
        extraction_tasks = []
        for field_name, prompt in field_prompts.items():
            task = {
                'field_name': field_name,
                'prompt': prompt,
                'master_class': created_classes[field_name]['master_class'],
            }
            extraction_tasks.append(task)
        
        print(f"Created {len(extraction_tasks)} extraction tasks")
        
        # 4. Define worker function for threaded extraction
        def extract_field_worker(task):
            """Worker function to extract a specific field type from document"""
            field_name = task['field_name']
            print(f"Starting extraction for field: {field_name}")
            
            try:
                response = self.client.responses.parse(
                    model=self.llm_model,
                    tools=[{
                        "type": "file_search",
                        "vector_store_ids": [self.vector_store_id],
                        "max_num_results": 2
                    }] if self.vector_store_id else None,
                    input=task['prompt'],
                    text_format=task['master_class'],
                )
                
                print(f"Successfully extracted field: {field_name}")
                print(f"Extracted data for {field_name}: {response.output_parsed}")
                
                return {
                    'field_name': field_name,
                    'success': True,
                    'extracted_data': response.output_parsed,
                    'error': None
                }
                
            except Exception as e:
                print(f"Error extracting field '{field_name}': {str(e)}")
                return {
                    'field_name': field_name,
                    'success': False,
                    'extracted_data': None,
                    'error': str(e)
                }
        
        # 5. Launch threads for parallel processing
        print(f"Starting parallel extraction with {len(extraction_tasks)} threads...")
        extraction_results = {}
        
        max_workers = min(len(extraction_tasks), 5)
        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            # Submit all tasks
            futures = []
            for task in extraction_tasks:
                future = executor.submit(extract_field_worker, task)
                futures.append(future)
            
            # Collect results as they complete
            for future in futures:
                result = future.result()
                field_name = result['field_name']
                extraction_results[field_name] = result
        
        print("Parallel extraction completed!")
        
        # 6. Compile final results (simplified structure)
        final_results = {}
        successful_extractions = 0
        failed_extractions = 0
        
        for field_name, extraction_result in extraction_results.items():
            final_results[field_name] = {
                'extraction_success': extraction_result['success'],
                'extracted_data': extraction_result['extracted_data'],
                'extraction_error': extraction_result['error'],
                'model_class': created_classes[field_name]['master_class']
            }
            
            if extraction_result['success']:
                successful_extractions += 1
            else:
                failed_extractions += 1
        
        print(f"Extraction summary: {successful_extractions} successful, {failed_extractions} failed")
        
        return final_results

    def process_multiple_sheets(self, sheets_dict: Dict[str, ListExcelFields]) -> Dict[str, List[Dict[str, Any]]]:
        """
        Process multiple sheets with their respective ExcelFields.
        Processes up to 4 sheets simultaneously to avoid API saturation.
        
        Args:
            sheets_dict: Dictionary mapping sheet names to ListExcelFields objects
            
        Returns:
            Dictionary mapping sheet names to lists of extracted field results
        """
        print(f"Processing {len(sheets_dict)} sheets...")
        
        # Log sheet information
        for sheet_name, excel_fields in sheets_dict.items():
            field_count = len(excel_fields.excelfields)
            print(f"  Sheet '{sheet_name}': {field_count} fields")
        
        # Define worker function for sheet-level processing
        def process_sheet_worker(sheet_item):
            """Worker function to process a single sheet"""
            sheet_name, list_excel_fields = sheet_item
            print(f"Starting processing for sheet: {sheet_name}")
            
            try:
                # Process all fields in this sheet
                sheet_results = self.process_dynamic_basemodels(list_excel_fields)
                
                # Convert results to the requested format
                extracted_fields = []
                for field_name, field_result in sheet_results.items():
                    extracted_field = {
                        'field_name': field_name,
                        'extraction_success': field_result['extraction_success'],
                        'extracted_data': field_result['extracted_data'],
                        'extraction_error': field_result['extraction_error']
                    }
                    extracted_fields.append(extracted_field)
                
                print(f"Successfully processed sheet: {sheet_name} ({len(extracted_fields)} fields)")
                
                return {
                    'sheet_name': sheet_name,
                    'success': True,
                    'extracted_fields': extracted_fields,
                    'error': None
                }
                
            except Exception as e:
                print(f"Error processing sheet '{sheet_name}': {str(e)}")
                return {
                    'sheet_name': sheet_name,
                    'success': False,
                    'extracted_fields': [],
                    'error': str(e)
                }
        
        # Process sheets in parallel (max 4 simultaneously)
        print(f"Starting parallel sheet processing with max 4 concurrent sheets...")
        sheet_results = {}
        
        max_workers = min(len(sheets_dict), 4)
        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            # Submit all sheet processing tasks
            futures = []
            for sheet_item in sheets_dict.items():
                future = executor.submit(process_sheet_worker, sheet_item)
                futures.append(future)
            
            # Collect results as they complete
            for future in futures:
                result = future.result()
                sheet_name = result['sheet_name']
                sheet_results[sheet_name] = result
        
        print("Parallel sheet processing completed!")
        
        # Compile final results in the requested format
        final_results = {}
        successful_sheets = 0
        failed_sheets = 0
        total_fields = 0
        successful_fields = 0
        
        for sheet_name, sheet_result in sheet_results.items():
            if sheet_result['success']:
                final_results[sheet_name] = sheet_result['extracted_fields']
                successful_sheets += 1
                
                # Count field-level successes
                for field in sheet_result['extracted_fields']:
                    total_fields += 1
                    if field['extraction_success']:
                        successful_fields += 1
            else:
                # Return empty list for failed sheets, but include error info
                final_results[sheet_name] = [{
                    'sheet_error': sheet_result['error'],
                    'extraction_success': False
                }]
                failed_sheets += 1
        
        print(f"\n=== PROCESSING SUMMARY ===")
        print(f"Sheets processed: {successful_sheets}/{len(sheets_dict)} successful")
        print(f"Fields processed: {successful_fields}/{total_fields} successful")
        print(f"Failed sheets: {failed_sheets}")
        
        return final_results

# Example usage
if __name__ == "__main__":
    # Initialize processor
    from openai import OpenAI
    from dotenv import load_dotenv
    from openpyxl import load_workbook
    
    load_dotenv()  # Load environment variables from .env file
    client = OpenAI()

    # Initialize DocumentProcessor
    processor = DocumentProcessor(
        client=client,
        llm_model="gpt-5-2025-08-07",  # Use actual OpenAI model name

    )
    # Initialize ExcelParser
    parser = ExcelParser(
        "Caledonia Facility A.xlsx",
        only_sheets=[load_workbook("Caledonia Facility A.xlsx").sheetnames[0]],
        openai_model="gpt-5-mini-2025-08-07",   # Use actual OpenAI model name
        lp_model="openai-gpt-4o-mini",          # the LlamaParse internal LLM
        max_concurrency=4,                      # control parallelism here
    )
    result, output = parser.parse_sheets()  # sync: wraps async
    multi_sheet_results = processor.process_multiple_sheets(result)

    i = 0
    # Process all sheets
    for sheet_name, fields in multi_sheet_results.items():
        print(f"\n=== SHEET: {sheet_name} ===")
        for field_result in fields:
            if field_result['extraction_success']:
                field_name = field_result['field_name']
                data = field_result['extracted_data']
                print(f"{field_name}: {data}")

                print(f"================")
                
                # Acceder a campos específicos
                if hasattr(data, 'full_name'):
                    print(f"  Client name: {data.full_name}")
                if hasattr(data, 'contract_value'):
                    print(f"  Contract value: {data.contract_value}")
    # Export results to Excel
    #print("\nExporting results to Excel...")
    #exporter = ExcelExporter()
    #output_file = exporter.export_to_excel(
        #results,  # From DocumentProcessor.process_multiple_sheets()
        #"contract_extraction_results.xlsx",
        #include_summary=True
    #)

    #print(f"Extraction results exported to: {output_file}")

    # Optional: Also use the convenience function
    # output_file = export_extraction_results(results, "contract_extraction_results.xlsx")