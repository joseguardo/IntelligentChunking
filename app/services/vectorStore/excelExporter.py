"""
Excel Export Pipeline for RAG Extraction Results

This module converts the nested Pydantic extraction results back into 
well-structured Excel files with intelligent data layout and basic formatting.
"""

from typing import Dict, List, Any, Optional, Union
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import re
from pathlib import Path


class ExcelExporter:
    """
    Converts RAG pipeline extraction results back to structured Excel files.
    
    Strategy: Create clean, well-structured Excel files that group related fields
    and handle variable-length extracted data intelligently.
    """
    
    def __init__(self):
        # Define styling constants
        self.header_font = Font(bold=True, color="FFFFFF")
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.subheader_font = Font(bold=True, color="000000")
        self.subheader_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'), 
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
    def normalize_extracted_data(self, extracted_data, expected_class=None):
        """
        Normalize extracted data to ensure consistent structure.
        Similar to normalize_model_response but for dynamically created classes.
        """
        if extracted_data is None:
            return None
            
        # If it's already the right type and has extractedfields, return as-is
        if hasattr(extracted_data, 'extractedfields'):
            return extracted_data
            
        # If it's a dict, try to convert it using the expected class
        if isinstance(extracted_data, dict) and expected_class:
            try:
                return expected_class(**extracted_data)
            except Exception as e:
                print(f"Warning: Could not normalize dict to {expected_class.__name__}: {e}")
                return None
        
        # If it's some other object, try to extract its dict representation
        if hasattr(extracted_data, 'model_dump'):
            try:
                data_dict = extracted_data.model_dump()
                if expected_class:
                    return expected_class(**data_dict)
            except Exception:
                pass
        elif hasattr(extracted_data, 'dict'):
            try:
                data_dict = extracted_data.dict()
                if expected_class:
                    return expected_class(**data_dict)
            except Exception:
                pass
        
        # If all else fails, return None (will be handled downstream)
        print(f"Warning: Could not normalize extracted_data of type {type(extracted_data)}")
        return None

    def sanitize_value_for_excel(self, value):
        """
        Sanitize values for Excel compatibility.
        
        TODO: TEMPORARY FIX - Convert datetime to string
        This should be reverted once we implement proper datetime handling
        that either:
        1. Strips timezone info (value.replace(tzinfo=None))
        2. Converts to Excel-compatible format
        3. Adds proper date formatting in Excel cells
        """
        from datetime import datetime, date, time
        
        # Handle datetime objects with timezone info
        if isinstance(value, (datetime, date, time)):
            # TEMPORARY: Convert to string to avoid Excel timezone error
            # TODO: Implement proper datetime handling
            if isinstance(value, datetime) and value.tzinfo is not None:
                # For now, convert to string format
                return value.strftime("%Y-%m-%d %H:%M:%S %Z").strip()
            elif isinstance(value, datetime):
                # Excel can handle timezone-naive datetime
                return value
            elif isinstance(value, date):
                return value
            elif isinstance(value, time):
                return value.strftime("%H:%M:%S")
        
        # Handle None values
        if value is None:
            return ""
        
        # Handle other types that might cause issues
        if hasattr(value, '__dict__') and not isinstance(value, (str, int, float, bool)):
            # Convert complex objects to string representation
            return str(value)
        
        return value

    def extract_values_from_pydantic(self, extracted_data, expected_class=None) -> List[Any]:
        """
        Extract actual values from the nested Pydantic structure.
        
        Args:
            extracted_data: The Pydantic model (e.g., ContactNameExtracted)
            expected_class: Optional expected class for normalization
            
        Returns:
            List of extracted values (sanitized for Excel compatibility)
        """
        # First, try to normalize the data
        normalized_data = self.normalize_extracted_data(extracted_data, expected_class)
        if not normalized_data:
            return []
            
        if not hasattr(normalized_data, 'extractedfields'):
            return []
            
        if not normalized_data.extractedfields:
            return []
            
        values = []
        for item in normalized_data.extractedfields:
            # Use model_fields to get only actual Pydantic fields, not methods
            if hasattr(item, 'model_fields'):
                for field_name in item.model_fields.keys():
                    if hasattr(item, field_name):
                        raw_value = getattr(item, field_name)
                        if raw_value is not None:
                            # SANITIZE VALUE FOR EXCEL COMPATIBILITY
                            sanitized_value = self.sanitize_value_for_excel(raw_value)
                            values.append(sanitized_value)
            else:
                # Fallback: try to get the actual field value more carefully
                # Convert the item to dict and extract non-None values
                try:
                    item_dict = item.model_dump() if hasattr(item, 'model_dump') else item.dict()
                    for key, raw_value in item_dict.items():
                        if raw_value is not None:
                            # SANITIZE VALUE FOR EXCEL COMPATIBILITY
                            sanitized_value = self.sanitize_value_for_excel(raw_value)
                            values.append(sanitized_value)
                except Exception:
                    # Last resort: try to find field values manually
                    for attr_name in dir(item):
                        if (not attr_name.startswith('_') and 
                            not callable(getattr(item, attr_name)) and
                            attr_name not in ['model_fields', 'model_config', 'model_computed_fields']):
                            raw_value = getattr(item, attr_name)
                            if raw_value is not None:
                                # SANITIZE VALUE FOR EXCEL COMPATIBILITY
                                sanitized_value = self.sanitize_value_for_excel(raw_value)
                                values.append(sanitized_value)
                        
        return values

    def group_related_fields(self, field_results: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """
        Group related fields together (e.g., Contact Name + Contact Email).
        
        Uses pattern matching to identify field relationships.
        """
        # Group fields by base name (remove numbers, common suffixes)
        groups = {}
        ungrouped = []
        
        for field_result in field_results:
            field_name = field_result['field_name']
            
            # Try to identify base pattern
            base_pattern = self._extract_base_pattern(field_name)
            
            if base_pattern and any(keyword in base_pattern.lower() for keyword in 
                                  ['contact', 'party', 'signatory', 'witness', 'vendor', 'client']):
                if base_pattern not in groups:
                    groups[base_pattern] = []
                groups[base_pattern].append(field_result)
            else:
                ungrouped.append(field_result)
        
        # Convert groups to structured format
        structured_groups = []
        
        # Add grouped fields
        for group_name, group_fields in groups.items():
            structured_groups.append({
                'type': 'group',
                'name': group_name,
                'fields': group_fields
            })
            
        # Add individual fields
        for field in ungrouped:
            structured_groups.append({
                'type': 'individual', 
                'name': field['field_name'],
                'fields': [field]
            })
            
        return structured_groups

    def _extract_base_pattern(self, field_name: str) -> str:
        """Extract base pattern from field name (e.g., 'Contact Name' from 'Contact 1 Name')"""
        # Remove numbers and common patterns
        cleaned = re.sub(r'\b\d+\b', '', field_name)  # Remove standalone numbers
        cleaned = re.sub(r'\s+', ' ', cleaned).strip()  # Clean whitespace
        return cleaned

    def create_summary_sheet(self, all_sheets_data: Dict[str, List[Dict[str, Any]]]) -> pd.DataFrame:
        """
        Create a summary sheet showing extraction statistics across all sheets.
        """
        summary_data = []
        
        for sheet_name, field_results in all_sheets_data.items():
            total_fields = len(field_results)
            successful_fields = sum(1 for f in field_results if f.get('extraction_success', False))
            total_values = 0
            
            for field_result in field_results:
                if field_result.get('extraction_success', False):
                    values = self.extract_values_from_pydantic(field_result.get('extracted_data'))
                    total_values += len(values)
            
            summary_data.append({
                'Sheet Name': sheet_name,
                'Total Fields': total_fields,
                'Successful Extractions': successful_fields,
                'Success Rate': f"{(successful_fields/total_fields*100):.1f}%" if total_fields > 0 else "0%",
                'Total Values Extracted': total_values
            })
        
        return pd.DataFrame(summary_data)

    def create_field_data_sheet(self, field_results: List[Dict[str, Any]], sheet_name: str) -> pd.DataFrame:
        """
        Create structured data sheet for a single original sheet's fields.
        """
        print(f"Creating data sheet for: {sheet_name}")
        
        # Group related fields
        grouped_fields = self.group_related_fields(field_results)
        print(f"  Grouped into {len(grouped_fields)} groups")
        
        # Build structured data
        rows = []
        
        for group in grouped_fields:
            if group['type'] == 'group':
                # Handle grouped fields (like contacts)
                try:
                    group_data = self._process_grouped_fields(group['fields'])
                    for row in group_data:
                        row['Group'] = group['name']
                        rows.append(row)
                    print(f"  Processed group '{group['name']}': {len(group_data)} rows")
                except Exception as e:
                    print(f"  Error processing group '{group['name']}': {e}")
                    # Add error row
                    rows.append({'Group': group['name'], 'Error': str(e)})
            else:
                # Handle individual fields
                try:
                    field_result = group['fields'][0]
                    field_data = self._process_individual_field(field_result)
                    for row in field_data:
                        row['Group'] = 'Individual Fields'
                        rows.append(row)
                    print(f"  Processed individual field '{group['name']}': {len(field_data)} rows")
                except Exception as e:
                    print(f"  Error processing individual field '{group['name']}': {e}")
                    # Add error row
                    rows.append({'Group': 'Individual Fields', 'Field': group['name'], 'Error': str(e)})
        
        if not rows:
            print(f"  No data extracted for {sheet_name}")
            return pd.DataFrame({'Message': ['No data extracted for this sheet']})
        
        print(f"  Total rows created: {len(rows)}")
        return pd.DataFrame(rows)

    def _process_grouped_fields(self, group_fields: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """Process a group of related fields (e.g., all contact-related fields)"""
        # Extract all values for each field in the group
        field_values = {}
        max_length = 0
        
        for field_result in group_fields:
            field_name = field_result['field_name']
            if field_result.get('extraction_success', False):
                extracted_data = field_result.get('extracted_data')
                expected_class = field_result.get('field_metadata', {}).get('master_class')
                values = self.extract_values_from_pydantic(extracted_data, expected_class)
                field_values[field_name] = values
                max_length = max(max_length, len(values))
            else:
                field_values[field_name] = []
        
        # Create rows by aligning related values
        rows = []
        for i in range(max(max_length, 1)):
            row = {'Index': i + 1}
            for field_name, values in field_values.items():
                row[field_name] = values[i] if i < len(values) else ""
            rows.append(row)
        
        return rows

    def _process_individual_field(self, field_result: Dict[str, Any]) -> List[Dict[str, Any]]:
        """Process an individual field"""
        field_name = field_result['field_name']
        
        if not field_result.get('extraction_success', False):
            return [{'Field': field_name, 'Value': 'EXTRACTION FAILED', 'Error': field_result.get('extraction_error', '')}]
        
        extracted_data = field_result.get('extracted_data')
        expected_class = field_result.get('field_metadata', {}).get('master_class')
        values = self.extract_values_from_pydantic(extracted_data, expected_class)
        
        if not values:
            return [{'Field': field_name, 'Value': 'NO DATA FOUND', 'Error': ''}]
        
        rows = []
        for i, value in enumerate(values):
            rows.append({
                'Field': field_name,
                'Value': value,
                'Index': i + 1 if len(values) > 1 else '',
                'Error': ''
            })
        
        return rows

    def apply_formatting(self, worksheet, df: pd.DataFrame):
        """Apply basic formatting to make the Excel sheet readable"""
        # Style headers
        for cell in worksheet[1]:
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.border
        
        # Style data cells and adjust column widths
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.border = self.border
                cell.alignment = Alignment(vertical='top', wrap_text=True)
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            worksheet.column_dimensions[column_letter].width = adjusted_width

    def export_to_excel(self, 
                       extraction_results: Dict[str, List[Dict[str, Any]]], 
                       output_path: Union[str, Path],
                       include_summary: bool = True) -> str:
        """
        Main export function that creates Excel file from extraction results.
        
        Args:
            extraction_results: Output from DocumentProcessor.process_multiple_sheets()
            output_path: Where to save the Excel file
            include_summary: Whether to include summary sheet
            
        Returns:
            Path to created Excel file
        """
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create summary sheet if requested
        if include_summary:
            summary_df = self.create_summary_sheet(extraction_results)
            summary_ws = wb.create_sheet("Summary")
            
            for row in dataframe_to_rows(summary_df, index=False, header=True):
                summary_ws.append(row)
            
            self.apply_formatting(summary_ws, summary_df)
        
        # Create sheet for each original sheet's data
        for sheet_name, field_results in extraction_results.items():
            # Clean sheet name for Excel compatibility
            clean_name = re.sub(r'[\\/*?:\[\]]', '_', sheet_name)[:31]
            
            field_data_df = self.create_field_data_sheet(field_results, sheet_name)
            data_ws = wb.create_sheet(f"{clean_name}_Data")
            
            for row in dataframe_to_rows(field_data_df, index=False, header=True):
                data_ws.append(row)
            
            self.apply_formatting(data_ws, field_data_df)
        
        # Save workbook
        output_path = Path(output_path)
        wb.save(output_path)
        
        return str(output_path)


# Integration function for the existing pipeline
def export_extraction_results(extraction_results: Dict[str, List[Dict[str, Any]]], 
                            output_filename: str = "extraction_results.xlsx") -> str:
    """
    Convenience function to export results from the RAG pipeline.
    
    Args:
        extraction_results: Output from DocumentProcessor.process_multiple_sheets()
        output_filename: Name for the output Excel file
        
    Returns:
        Path to created Excel file
    """
    exporter = ExcelExporter()
    return exporter.export_to_excel(extraction_results, output_filename)


