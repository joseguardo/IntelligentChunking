# The goal of this modules is to create a pipeline that: 
# 1. Takes an excel template file as input and sends it to the OpenAI API one sheet at a time.
# 2. Divides it into the different sheets that it contains. 
# 3. Processes each sheet in order to find all of the fields of the template that need to be completed in the sheet. 
# 4. For each field it will generate an entry in a BaseModel where it will discribe the name of the field (which must be the same as in the excel tempplate), brief description, and the type of data that the template requires for the field. 
# 5. Finally it will return a list of BaseModels, one for each sheet in the excel template.


# GOAL: Automatically transform excel templates into structured data models that allow autonomized search and retrieval of information from filled templates.

import os
import re
import tempfile
import asyncio
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional
from textwrap import dedent
from pprint import pprint


from pydantic import BaseModel, Field
from openpyxl import load_workbook, Workbook
from dotenv import load_dotenv
from openai import OpenAI  # pip install openai
from llama_cloud_services import LlamaParse  # pip install llama-cloud-services (per docs)

# ---------- Your Pydantic models ----------
class ExcelField(BaseModel):
    name: Optional[str] = Field(None, description="The name of the field as it appears in the excel template.")
    description: Optional[str] = Field(None, description="A brief description of the field.")
    data_type: Optional[str] = Field(None, description="The type of data that the field requires. E.g., 'str', 'int', 'float', 'datetime'.")
    prompt: Optional[str] = Field(None, description="A prompt that can be used to extract the field value from the contract and that contains context from the sheet and other fields present.")

class ListExcelFields(BaseModel):
    excelfields: List[ExcelField] = Field(..., description="A list of fields extracted from the excel template.")

# ---------- Helpers ----------
SAFE_SHEET_CHARS = r'[\\/*?:\[\]]'
def _sanitize_sheet_name(name: str, maxlen: int = 31) -> str:
    n = re.sub(SAFE_SHEET_CHARS, "_", (name or "").strip())
    return n[:maxlen] or "Sheet"

def _write_single_sheet_tempfile(src_ws) -> str:
    """
    Create a new xlsx on disk containing only the provided worksheet's VALUES (no styles).
    Returns path to the temp file.
    """
    tmp = tempfile.NamedTemporaryFile(prefix="sheet_", suffix=".xlsx", delete=False)
    tmp_path = tmp.name
    tmp.close()

    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = _sanitize_sheet_name(src_ws.title)

    for row in src_ws.iter_rows(values_only=True):
        ws_out.append(row)

    # (optional) column widths for readability (best-effort)
    try:
        for dim, col_dim in src_ws.column_dimensions.items():
            if col_dim.width:
                ws_out.column_dimensions[dim].width = col_dim.width
    except Exception:
        pass

    wb_out.save(tmp_path)
    return tmp_path

# ---------- Main parser ----------
class ExcelParser:
    @staticmethod
    def split_listexcelfields(parsed: 'ListExcelFields') -> list:
        """
        Given a ListExcelFields object, return a list of ListExcelFields,
        each containing a single ExcelField from the original.
        """
        return [ListExcelFields(excelfields=[field]) for field in (parsed.excelfields or [])]
    def __init__(
        self,
        excel_file_path: str,
        *,
        openai_model: Optional[str] = None,       # model for responses.parse
        include_hidden: bool = False,
        only_sheets: Optional[List[str]] = None,
        exclude_sheets: Optional[List[str]] = None,
        max_concurrency: int = 8,  # CHANGE: used to cap concurrent sheet processing

        # LlamaParse config (matches your snippet; override if desired)
        lp_parse_mode: str = "parse_page_with_agent",
        lp_model: str = "openai-gpt-4o-mini",
        lp_high_res_ocr: bool = True,
        lp_adaptive_long_table: bool = True,
        lp_outlined_table_extraction: bool = True,
        lp_output_tables_as_html: bool = True,
    ):
        """
        We still return Dict[str, ListExcelFields] to keep your public API intact.
        """
        load_dotenv()

        self.excel_file_path = str(Path(excel_file_path).resolve())
        self.workbook = load_workbook(self.excel_file_path, data_only=True)
        self.max_concurrency = max(1, int(max_concurrency))  # CHANGE: store cap

        # OpenAI client for the final structured extraction
        self.client = OpenAI(api_key=os.getenv("OPENAI_API_KEY"))
        self.openai_model = openai_model or os.getenv("OPENAI_MODEL") or "gpt-4.1"

        # LlamaParse client for robust parsing of each sheet temp file
        self.llama_parser = LlamaParse(
            api_key=os.getenv("LLAMAPARSE_API_KEY"),
            parse_mode=lp_parse_mode,
            model=lp_model,
            high_res_ocr=lp_high_res_ocr,
            adaptive_long_table=lp_adaptive_long_table,
            outlined_table_extraction=lp_outlined_table_extraction,
            output_tables_as_HTML=lp_output_tables_as_html,
        )

        # choose sheets
        all_sheet_names = self.workbook.sheetnames
        if only_sheets:
            target = [s for s in all_sheet_names if s in set(only_sheets)]
        else:
            target = list(all_sheet_names)
        if exclude_sheets:
            target = [s for s in target if s not in set(exclude_sheets)]
        if not include_hidden:
            visible = []
            for s in target:
                ws = self.workbook[s]
                if getattr(ws, "sheet_state", "visible") == "hidden":
                    continue
                visible.append(s)
            target = visible
        self.sheets = target

    # ----------------- Public API (sync) -----------------
    def parse_sheets(self) -> Dict[str, ListExcelFields]:
        """
        Sync wrapper: internally runs async LlamaParse per sheet, then OpenAI parse.
        """
        return asyncio.run(self.parse_sheets_async())

    # ----------------- Async implementation -----------------
    async def parse_sheets_async(self) -> Dict[str, ListExcelFields]:
        results: Dict[str, ListExcelFields] = {}
        sem = asyncio.Semaphore(self.max_concurrency)

        # create tasks
        tasks = [self._process_sheet(sheet_name, sem) for sheet_name in self.sheets]

        # print a header so it's visually clear
        print(f"[{datetime.now():%H:%M:%S}] 🚀 Starting {len(tasks)} sheets with max_concurrency={self.max_concurrency}")
        output_dict = {}
        # show results as they COMPLETE (not in submission order)
        for fut in asyncio.as_completed(tasks):
            sheet_name, parsed = await fut
            results[sheet_name] = parsed

            # live, per-sheet preview
            try:
                line = self._preview_line(sheet_name, parsed, getattr(self, "preview_n", 5))  # CHANGE
                print(line)
            except Exception:
                # don't let logging crash the run
                print(f"[{datetime.now():%H:%M:%S}] ✅ {sheet_name}: {len(parsed.excelfields)} fields")

            try:
                parsed = self.split_listexcelfields(parsed)  # CHANGE: test the new static method
                output_dict[sheet_name] = parsed
            except Exception:
                pass

        print(f"[{datetime.now():%H:%M:%S}] 🏁 All sheets done.")
        return results, output_dict


    # CHANGE: new async per-sheet pipeline (guarded by semaphore)
    async def _process_sheet(self, sheet_name: str, sem: asyncio.Semaphore):
        await sem.acquire()
        tmp_path = None
        try:
            ws = self.workbook[sheet_name]
            tmp_path = _write_single_sheet_tempfile(ws)

            # 1) LlamaParse (async)
            parsed_text = await self._llamaparse_sheet_text(tmp_path)

            # 2) OpenAI structuring (sync client -> run in thread)
            result = await self._openai_extract_fields_from_text(sheet_name, parsed_text)

            # Normalize
            result = self.normalize_model_response(result)
            orig_len = len(result.excelfields or [])
            orig_fields = [f.name for f in (result.excelfields or []) if getattr(f, "name", None)]

            # 3) OpenAI validation: iterate through fields and look for extracted fields that could be part of a list such as "Contact 1 Name", "Contact 2 Name", etc.
            result = await self._openai_validate_fields(sheet_name, result)

            # Normalize
            result = self.normalize_model_response(result)
            new_fields = [f.name for f in (result.excelfields or []) if getattr(f, "name", None)]
            new_len = len(result.excelfields or [])
            if new_len != orig_len:
                print(f"[{datetime.now():%H:%M:%S}] 🔄 {sheet_name}: Validated fields, reduced {orig_len} → {new_len}")
                print(f"  - Removed fields: {set(orig_fields) - set(new_fields)}")
                print(f"  - Added fields: {set(new_fields) - set(orig_fields)}")


            return sheet_name, result

        except Exception as e:
            print(f"[ERROR] Sheet '{sheet_name}': {e}")
            return sheet_name, ListExcelFields(excelfields=[])
        finally:
            try:
                if tmp_path and os.path.exists(tmp_path):
                    os.remove(tmp_path)
            except Exception:
                pass
            sem.release()

    # ----------------- Internals -----------------
    async def _llamaparse_sheet_text(self, file_path: str) -> str:
        """
        Use LlamaParse to parse a single-sheet XLSX. Return a single text string.
        We concatenate all returned page texts (if any).
        """
        start = datetime.now()
        print(f"[LlamaParse] Parsing: {Path(file_path).name} at {start:%Y-%m-%d %H:%M:%S}")

        # The docs show:
        # result = await parser.aparse("dcf_template.xlsx")
        # llama_parse_documents = result.get_text_documents(split_by_page=True)
        result = await self.llama_parser.aparse(file_path)
        docs = result.get_text_documents(split_by_page=True)

        # You can choose HTML or text—docs[i].text should contain the parsed content.
        # If you set output_tables_as_HTML=True, text may include HTML tables.
        combined = "\n\n".join([d.text for d in docs if getattr(d, "text", None)])
        if not combined.strip():
            print("[LlamaParse] Warning: Empty parse result; falling back to empty text.")
        return combined

    # CHANGE: split into sync core (runs OpenAI client) ...
    def _openai_extract_fields_from_text_sync(self, sheet_name: str, parsed_text: str) -> ListExcelFields:
        """
        Call OpenAI Responses API to convert the LlamaParse text into your Pydantic structure. (SYNC)
        """
        prompt = dedent("""
            You are given the content of an EXCEL SHEET TEMPLATE that has been parsed into text/HTML by an OCR/structure parser.
            Your task is to identify any explicit fillable fields in the template.

            For each field you find, return:
            1) The exact field name as it appears in the template.
            2) A brief description of the field.
            3) The expected data type from: 'str', 'int', 'float', 'datetime'.
            4) A prompt that captures the context of the field through the purpose of the sheet and enables the extraction of the field value from a contract through a vector store search.

            Important rules:
            - ONLY include fields explicitly present in the sheet.
            - Understand the template structure and differentiate between labels, headers, and actual fields to fill.
            - DO NOT INCLUDE HEADERS, TITLES, OR ANYTHING NOT A FIELD: The idea is to extract only fields that need to be filled, not to change the template. 
            - Do not hallucinate.
            - Ignore generic headers, boilerplate, or instructional text unless it is itself a field to fill.
            - If the parser emitted tables as HTML, read them as if they were the original grid.

            Return strictly as the Pydantic model: ListExcelFields.
        """).strip()

        try:
            start = datetime.now()
            print(f"[OpenAI] Structuring fields for '{sheet_name}' at {start:%H:%M:%S}")

            completion = self.client.responses.parse(
                model=self.openai_model,
                input=[{
                    "role": "user",
                    "content": [
                        {"type": "input_text", "text": parsed_text[:200000]},  # safety trim if huge
                        {"type": "input_text", "text": prompt},
                    ],
                }],
                text_format=ListExcelFields,
            )
            structured = completion.output_parsed
            print(f"[OpenAI] Done '{sheet_name}' with {len(structured.excelfields)} fields.")
            return structured

        except Exception as e:
            print(f"[OpenAI ERROR] '{sheet_name}': {e}")
            return ListExcelFields(excelfields=[])
        


    def _openai_validate_fields_sync(self, sheet_name: str, result: ListExcelFields) -> ListExcelFields:
        """ 
        Call the OpenAI Api to validate the fields extracted from the excel sheet. (SYNC)
        We look for patterns in the field names that suggest they are part of a list (e.g., "Contact 1 Name", "Contact 2 Name").
        If we find such patterns, we consolidate them into a single field. Since the ListExcelFields already is formatted as a list, it would be as if we removed duplicates. 
        """

        prompt = dedent(""" 
        You are an expert data structuring assistant specializing in the normalization of extracted form fields from Excel templates.
        You are given a list of extracted fields, each with a name, description, and data type, as parsed from an Excel sheet.
        Your task is to analyze the list of field names and identify patterns that indicate the fields are part of a repeated group or list.
        For example, fields such as "Contact 1 Name", "Contact 2 Name", "Contact 3 Name" should be recognized as instances of a single logical field ("Contact Name") that occurs multiple times.

        Instructions:

        - Review all field names for numeric or sequential patterns (e.g., "Field 1", "Field 2", "Field 3", etc.).
        - For each group of fields that share a common base name and differ only by a number or similar index, consolidate them into a single field definition.
        - In the output, ensure that each logical field appears only once, with a clear indication in the description that it can occur multiple times if applicable.
        - Data type should still only be one of: 'str', 'int', 'float', 'datetime'.
        - If no patterns are found, return the original list unchanged.
        - Ensure the final output is a valid ListExcelFields Pydantic model.
        - Be careful with fields that may contain same words, but are not part of a list (e.g., "Start Date" and "End Date" are distinct and should not be consolidated).
        - Preserve all unique, non-repeated fields as-is.
        - Return the validated and consolidated list strictly as a ListExcelFields Pydantic model, ensuring the output is clean, deduplicated, and semantically accurate.

        Example: 
            Input: 
            ListExcelFields(excelfields=[
            ExcelField(name="Contact 1 Name", description="Name of the first contact", data_type="str", prompt = "..."),
            ExcelField(name="Contact 2 Name", description="Name of the second contact", data_type="str", prompt = "..."),
            ExcelField(name="Contact 1 Email", description="Email of the first contact", data_type="str", prompt = "..."),
            ExcelField(name="Contact 2 Email", description="Email of the second contact", data_type="str", prompt = "..."),
            ExcelField(name="Company Name", description="Name of the company", data_type="str", prompt = "..."),
        ])
        Output: 
                    ListExcelFields(excelfields=[
                ExcelField(
                    name="Contact Name",
                    description="Name of each contact (list, one per contact)",
                    data_type="str", 
                    prompt = "Extract the name of each contact from the contract."
                ),
                ExcelField(
                    name="Contact Email",
                    description="Email of each contact (list, one per contact)",
                    data_type="str",
                    prompt = "Extract the email of each contact from the contract."
                ),
                ExcelField(
                    name="Company Name",
                    description="Name of the company",
                    data_type="str",
                    prompt = "Extract the name of the company from the contract."
                ),
            ])
        Do not hallucinate or invent fields. Only consolidate when a clear pattern is present. If no patterns are found, return the original list unchanged.
        """).strip()
        try:
            start = datetime.now()
            print(f"[OpenAI] Validating fields for '{sheet_name}' at {start:%H:%M:%S}")

            # Convert the ListExcelFields to a string representation for the prompt
            fields_str = repr(result)

            completion = self.client.responses.parse(
                model=self.openai_model,
                input=[{
                    "role": "user",
                    "content": [
                        {"type": "input_text", "text": fields_str+prompt},
                    ],
                }],
                text_format=ListExcelFields,
            )
            validated = completion.output_parsed
            print(f"[OpenAI] Validation done '{sheet_name}' with {len(validated.excelfields)} fields.")
            return validated

        except Exception as e:
            print(f"[OpenAI ERROR] Validation '{sheet_name}': {e}")
            return result  # return original if validation fails
        


    # CHANGE: ... and async wrapper so we can run it concurrently without blocking the loop
    async def _openai_extract_fields_from_text(self, sheet_name: str, parsed_text: str) -> ListExcelFields:
        return await asyncio.to_thread(self._openai_extract_fields_from_text_sync, sheet_name, parsed_text)
    
    async def _openai_validate_fields(self, sheet_name: str, result: ListExcelFields) -> ListExcelFields:
        return await asyncio.to_thread(self._openai_validate_fields_sync, sheet_name, result)
    
    # CHANGE: helper to build a compact preview line
    def _preview_line(self, sheet_name: str, parsed: ListExcelFields, n: int = 5) -> str:
        names = [f.name for f in (parsed.excelfields or []) if getattr(f, "name", None)]
        head = ", ".join(names) if names else "—"
        #more = f" +{len(names)-n}" if len(names) > n else ""
        return f"[{datetime.now():%H:%M:%S}] ✅ {sheet_name}: {len(parsed.excelfields)} fields  →  {head}"
    
    def normalize_model_response(self, response) -> ListExcelFields:
        if isinstance(response, dict):
            return ListExcelFields(**response)
        elif isinstance(response, ListExcelFields):
            return response
        else:
            return ListExcelFields(excelfields=[])
        
    def convert_excelfield_to_listexcelfields(self, field: ExcelField) -> ListExcelFields:
        return ListExcelFields(excelfields=[field])


if __name__ == "__main__":
    # env:
    #   OPENAI_API_KEY=...
    #   LLAMAPARSE_API_KEY=...
    #   (optional) OPENAI_MODEL=gpt-4.1
    parser = ExcelParser(
        "Caledonia Facility A.xlsx",
        only_sheets=[load_workbook("Caledonia Facility A.xlsx").sheetnames[0]],
        openai_model="gpt-5-mini-2025-08-07",   # the structuring step
        lp_model="openai-gpt-4o-mini",          # the LlamaParse internal LLM (per your snippet)
        max_concurrency=4,                      # CHANGE: control parallelism here
    )
    result, output = parser.parse_sheets()  # sync: wraps async
    pprint(result)
